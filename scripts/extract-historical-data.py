"""
Extrae la data histórica del Excel `vault/inbox/2026-05-17-historico-envases.xlsx`
y la convierte en `src/lib/data/historical-data.json` para alimentar el mock del
dashboard con datos reales.

Decisiones (acordadas con el usuario el 2026-05-17):
- Solo se incluye Airkem (los 4 registros con typos "ION - Airkem", "Airkem-ION",
  "Handy Solutions" se descartan).
- Se generan 189 containers `A-001 .. A-189` con la tara observada en el Excel
  (tara es constante por carro). Todos son 240L (el Excel se llama "Carro 240 L Nº").
- waste_type = 'infectious' para todos (no viene en el Excel; es el tipo más común
  y el que se trata on-site en autoclave).
- Pesos netos <= 0 se descartan (errores de captura).
- Treatment runs:
  * Si en el Excel `Fecha de Tratado == Fecha de pesaje`: se simula un rezago de
    4 horas tras el pesaje (tiempo típico de autoclave on-site).
  * Si `Fecha de Tratado > Fecha de pesaje`: se respeta esa fecha y se asigna
    hora 07:00 (amanecer del día siguiente).
- Storage events: entry 5 min después del pesaje, exit = inicio del tratamiento.
- Weighing sessions: 1 por día con todas las recepciones del día.
- Operadores: se rota entre `user-1` y `user-2`.
- IDs cortos (`r{n}`, `s{n}`, `t{n}`, `ws{n}`) para mantener el JSON liviano.
"""

from __future__ import annotations

import json
import re
import warnings
from collections import defaultdict
from datetime import datetime, timedelta, time
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parent.parent
EXCEL = ROOT / "vault" / "inbox" / "2026-05-17-historico-envases.xlsx"
OUTPUT = ROOT / "src" / "lib" / "data" / "historical-data.json"

# Zona horaria de Panamá: UTC-5 (sin DST).
TZ_OFFSET = "-05:00"

COMPANY_ID = "company-airkem"
COMPANY_LETTER = "A"
OPERATORS = ["user-1", "user-2"]


def parse_time(hora_str: str) -> time:
    """'06:04:55-AM' -> time(6, 4, 55). '12:30:00-PM' -> time(12, 30, 0)."""
    m = re.match(r"^(\d{1,2}):(\d{2}):(\d{2})-(AM|PM)$", str(hora_str).strip())
    if not m:
        raise ValueError(f"Hora no parseable: {hora_str!r}")
    h, mm, ss, ampm = int(m[1]), int(m[2]), int(m[3]), m[4]
    if ampm == "AM":
        if h == 12:
            h = 0
    else:
        if h != 12:
            h += 12
    return time(h, mm, ss)


def iso(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%dT%H:%M:%S") + TZ_OFFSET


def main() -> None:
    print(f"Leyendo {EXCEL.relative_to(ROOT)}...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb["Kilos Diarios"]

    # ── Pasada 1: agregar tara por carro y descubrir días ─────────────────────
    tara_por_carro: dict[int, float] = {}
    raw_rows: list[dict] = []

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        empresa, carro, fecha, hora, tara, peso, neto, obs, fecha_trat, *_ = row
        if empresa != "Airkem":
            continue
        if carro is None or not isinstance(carro, (int, float)):
            continue
        if fecha is None or hora is None or peso is None:
            continue
        if neto is None or neto <= 0:
            continue

        carro_num = int(carro)
        # La tara es constante por carro en el Excel; mantenemos el primer valor
        # observado (mediana == promedio == cualquier valor).
        if tara is not None and carro_num not in tara_por_carro:
            tara_por_carro[carro_num] = float(tara)

        raw_rows.append({
            "carro": carro_num,
            "fecha": fecha.date(),
            "hora_str": str(hora),
            "tara": float(tara) if tara is not None else 0.0,
            "peso": float(peso),
            "neto": float(neto),
            "obs": str(obs) if obs else None,
            "fecha_trat": fecha_trat.date() if fecha_trat else None,
        })

    print(f"  {len(raw_rows)} recepciones válidas, {len(tara_por_carro)} carros distintos")

    # Ordenar por fecha + hora para asignar IDs incrementales en orden cronológico
    raw_rows.sort(key=lambda r: (r["fecha"], parse_time(r["hora_str"])))

    # ── Containers (189) ──────────────────────────────────────────────────────
    containers = []
    for carro_num in sorted(tara_por_carro.keys()):
        cid = f"{COMPANY_LETTER}-{carro_num:03d}"
        containers.append({
            "id": cid,
            "company_id": COMPANY_ID,
            "size_liters": 240,
            "tare_weight_kg": round(tara_por_carro[carro_num], 2),
            "waste_type": "infectious",
            "status": "active",
            "registered_at": "2026-01-01T00:00:00" + TZ_OFFSET,
        })

    # ── Receptions, storage, treatment, weighing sessions ─────────────────────
    receptions = []
    storage_events = []
    treatment_runs = []

    # Una weighing session por día (agrupa todas las recepciones de ese día).
    sessions_by_day: dict[str, dict] = {}

    # Para generar la ubicación "regresó al cliente" tras el último tratamiento.
    last_treatment_end_by_container: dict[str, datetime] = {}

    for idx, r in enumerate(raw_rows, start=1):
        container_id = f"{COMPANY_LETTER}-{r['carro']:03d}"
        dt_pesaje = datetime.combine(r["fecha"], parse_time(r["hora_str"]))
        dt_entry_storage = dt_pesaje + timedelta(minutes=5)
        if r["fecha_trat"] and r["fecha_trat"] > r["fecha"]:
            # Tratado al día siguiente: 07:00 AM
            dt_treatment_start = datetime.combine(r["fecha_trat"], time(7, 0, 0))
        else:
            # Tratado el mismo día: 4 horas después del pesaje
            dt_treatment_start = dt_pesaje + timedelta(hours=4)
        dt_treatment_end = dt_treatment_start + timedelta(hours=1, minutes=30)

        op = OPERATORS[idx % len(OPERATORS)]
        day_iso = r["fecha"].isoformat()

        # Weighing session lazy-created per day
        if day_iso not in sessions_by_day:
            sessions_by_day[day_iso] = {
                "id": f"ws-{day_iso}",
                "client_id": "client-1",
                "date": day_iso,
                "started_at": iso(datetime.combine(r["fecha"], time(6, 0, 0))),
                "ended_at": iso(datetime.combine(r["fecha"], time(11, 0, 0))),
                "operator_id": op,
                "status": "completed",
                "reception_ids": [],
            }
        sessions_by_day[day_iso]["reception_ids"].append(f"r{idx}")

        receptions.append({
            "id": f"r{idx}",
            "container_id": container_id,
            "weighing_session_id": f"ws-{day_iso}",
            "arrived_at": iso(dt_pesaje),
            "gross_weight_kg": round(r["peso"], 2),
            "operator_id": op,
            "photo_ids": [],
        })

        storage_events.append({
            "id": f"s{idx}",
            "container_id": container_id,
            "entry_at": iso(dt_entry_storage),
            "exit_at": iso(dt_treatment_start),
            "operator_id": op,
            "photo_ids": [],
        })

        treatment_runs.append({
            "id": f"t{idx}",
            "container_id": container_id,
            "started_at": iso(dt_treatment_start),
            "completed_at": iso(dt_treatment_end),
            "operator_id": op,
        })

        # Para cada container, recordamos el último tratamiento finalizado.
        prev = last_treatment_end_by_container.get(container_id)
        if prev is None or dt_treatment_end > prev:
            last_treatment_end_by_container[container_id] = dt_treatment_end

    # Locations: tras el último tratamiento de cada container, el envase
    # regresa limpio al cliente. Esto evita que la torta de circulación los
    # clasifique a todos como "sin_registro".
    locations = []
    AREAS = [
        ("1", "Emergencias"),
        ("2", "Pediatría"),
        ("3", "UCI"),
        ("4", "Oncología"),
    ]
    for li, (container_id, dt_end) in enumerate(sorted(last_treatment_end_by_container.items()), start=1):
        floor, area = AREAS[li % len(AREAS)]
        locations.append({
            "id": f"loc-h{li}",
            "container_id": container_id,
            "reported_at": iso(dt_end + timedelta(hours=2)),
            "operator_id": OPERATORS[li % len(OPERATORS)],
            "location_type": "client_site",
            "client_id": "client-1",
            "floor": floor,
            "area": area,
            "notes": None,
        })

    weighing_sessions = list(sessions_by_day.values())
    weighing_sessions.sort(key=lambda s: s["date"])

    # ── Resumen ───────────────────────────────────────────────────────────────
    total_kg = sum(rec["gross_weight_kg"] - tara_por_carro[int(rec["container_id"][2:])]
                   for rec in receptions)
    print("\nResumen generado:")
    print(f"  containers:         {len(containers)}")
    print(f"  receptions:         {len(receptions)}")
    print(f"  storage_events:     {len(storage_events)}")
    print(f"  treatment_runs:     {len(treatment_runs)}")
    print(f"  weighing_sessions:  {len(weighing_sessions)}")
    print(f"  locations:          {len(locations)}")
    print(f"  total kg netos:     {total_kg:,.1f}")

    # ── Persistir JSON ────────────────────────────────────────────────────────
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "_meta": {
            "source": "vault/inbox/2026-05-17-historico-envases.xlsx",
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "date_range": {
                "from": raw_rows[0]["fecha"].isoformat(),
                "to": raw_rows[-1]["fecha"].isoformat(),
            },
        },
        "containers": containers,
        "weighing_sessions": weighing_sessions,
        "receptions": receptions,
        "storage_events": storage_events,
        "treatment_runs": treatment_runs,
        "locations": locations,
    }
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    size_mb = OUTPUT.stat().st_size / 1024 / 1024
    print(f"\nEscrito: {OUTPUT.relative_to(ROOT)}  ({size_mb:.2f} MB)")


if __name__ == "__main__":
    main()
